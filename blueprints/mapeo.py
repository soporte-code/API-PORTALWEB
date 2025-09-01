from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from utils.db import get_db_connection
import logging
import uuid
from datetime import datetime
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

mapeo_bp = Blueprint('mapeo_bp', __name__)
logger = logging.getLogger(__name__)

# ============================================================================
# REGISTROS DE MAPEO
# ============================================================================

# Obtener todos los registros de mapeo del usuario logueado
@mapeo_bp.route('/registros-mapeo', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_registros_mapeo():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener registros de mapeo de cuarteles accesibles al usuario
        cursor.execute("""
            SELECT rm.*, c.nombre as cuartel_nombre, c.id_ceco
            FROM mapeo_fact_registromapeo rm
            JOIN general_dim_cuartel c ON rm.id_cuartel = c.id
            WHERE c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
            ORDER BY rm.fecha_creacion DESC
        """, (usuario_id,))
        
        registros = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(registros), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo registros de mapeo: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Obtener un registro de mapeo específico
@mapeo_bp.route('/registros-mapeo/<string:registro_id>', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_registro_mapeo(registro_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al registro de mapeo
        cursor.execute("""
            SELECT rm.*, c.nombre as cuartel_nombre, c.id_ceco
            FROM mapeo_fact_registromapeo rm
            JOIN general_dim_cuartel c ON rm.id_cuartel = c.id
            WHERE rm.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
        """, (registro_id, usuario_id))
        
        registro = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not registro:
            return jsonify({"error": "Registro de mapeo no encontrado"}), 404
            
        return jsonify(registro), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo registro de mapeo {registro_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Crear un nuevo registro de mapeo
@mapeo_bp.route('/registros-mapeo', methods=['POST', 'OPTIONS'])
@jwt_required()
def crear_registro_mapeo():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['id_cuartel', 'fecha_inicio', 'fecha_termino']
        for campo in campos_requeridos:
            if campo not in data:
                return jsonify({"error": f"Campo requerido: {campo}"}), 400
        
        # Validar fechas
        try:
            fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            fecha_termino = datetime.strptime(data['fecha_termino'], '%Y-%m-%d').date()
            
            if fecha_inicio > fecha_termino:
                return jsonify({"error": "La fecha de inicio debe ser anterior a la fecha de término"}), 400
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al cuartel
        cursor.execute("""
            SELECT 1 FROM general_dim_cuartel c
            WHERE c.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
        """, (data['id_cuartel'], usuario_id))
        
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"error": "No tienes acceso a este cuartel"}), 403
        
        # Verificar que no exista un registro de mapeo activo para el cuartel
        cursor.execute("""
            SELECT 1 FROM mapeo_fact_registromapeo 
            WHERE id_cuartel = %s AND id_estado = 1
        """, (data['id_cuartel'],))
        
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"error": "Ya existe un registro de mapeo activo para este cuartel"}), 400
        
        # Generar UUID para el registro
        registro_id = str(uuid.uuid4())
        
        # Insertar el registro de mapeo
        cursor.execute("""
            INSERT INTO mapeo_fact_registromapeo (
                id, id_temporada, id_cuartel, fecha_inicio, fecha_termino, 
                id_estado, fecha_creacion
            ) VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (
            registro_id,
            data.get('id_temporada'),
            data['id_cuartel'],
            fecha_inicio,
            fecha_termino,
            1  # id_estado activo
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            "message": "Registro de mapeo creado exitosamente",
            "id": registro_id
        }), 201
        
    except Exception as e:
        logger.error(f"Error creando registro de mapeo: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Actualizar un registro de mapeo
@mapeo_bp.route('/registros-mapeo/<string:registro_id>', methods=['PUT', 'OPTIONS'])
@jwt_required()
def actualizar_registro_mapeo(registro_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al registro de mapeo
        cursor.execute("""
            SELECT 1 FROM mapeo_fact_registromapeo rm
            JOIN general_dim_cuartel c ON rm.id_cuartel = c.id
            WHERE rm.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
        """, (registro_id, usuario_id))
        
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"error": "Registro de mapeo no encontrado o sin acceso"}), 404
        
        # Validar fechas si se proporcionan
        if 'fecha_inicio' in data and 'fecha_termino' in data:
            try:
                fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
                fecha_termino = datetime.strptime(data['fecha_termino'], '%Y-%m-%d').date()
                
                if fecha_inicio > fecha_termino:
                    return jsonify({"error": "La fecha de inicio debe ser anterior a la fecha de término"}), 400
            except ValueError:
                return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}), 400
        
        # Actualizar el registro de mapeo
        campos_actualizables = ['id_temporada', 'fecha_inicio', 'fecha_termino', 'id_estado']
        set_clause = []
        valores = []
        
        for campo in campos_actualizables:
            if campo in data:
                set_clause.append(f"{campo} = %s")
                valores.append(data[campo])
        
        if not set_clause:
            cursor.close()
            conn.close()
            return jsonify({"error": "No hay campos para actualizar"}), 400
        
        valores.append(registro_id)
        
        cursor.execute(f"""
            UPDATE mapeo_fact_registromapeo 
            SET {', '.join(set_clause)}
            WHERE id = %s
        """, valores)
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"message": "Registro de mapeo actualizado exitosamente"}), 200
        
    except Exception as e:
        logger.error(f"Error actualizando registro de mapeo {registro_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# ============================================================================
# ESTADOS DE HILERAS
# ============================================================================

# Obtener estados de hileras de un registro de mapeo
@mapeo_bp.route('/registros-mapeo/<string:registro_id>/estados-hileras', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_estados_hileras(registro_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar acceso al registro de mapeo y obtener estados de hileras
        cursor.execute("""
            SELECT eh.*, h.hilera, c.nombre as cuartel_nombre, u.nombre as usuario_nombre
            FROM mapeo_fact_estado_hilera eh
            JOIN mapeo_fact_registromapeo rm ON eh.id_registro_mapeo = rm.id
            JOIN general_dim_hilera h ON eh.id_hilera = h.id
            JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_usuario u ON eh.id_usuario = u.id
            WHERE rm.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
            ORDER BY h.hilera ASC
        """, (registro_id, usuario_id))
        
        estados = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(estados), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo estados de hileras para registro {registro_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Crear/actualizar estado de hilera
@mapeo_bp.route('/registros-mapeo/<string:registro_id>/estados-hileras', methods=['POST', 'OPTIONS'])
@jwt_required()
def crear_estado_hilera(registro_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['id_hilera', 'estado']
        for campo in campos_requeridos:
            if campo not in data:
                return jsonify({"error": f"Campo requerido: {campo}"}), 400
        
        # Validar estado
        estados_validos = ['pendiente', 'en_progreso', 'pausado', 'completado']
        if data['estado'] not in estados_validos:
            return jsonify({"error": f"Estado inválido. Estados válidos: {', '.join(estados_validos)}"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al registro de mapeo y la hilera
        cursor.execute("""
            SELECT 1 FROM mapeo_fact_registromapeo rm
            JOIN general_dim_cuartel c ON rm.id_cuartel = c.id
            JOIN general_dim_hilera h ON h.id_cuartel = c.id
            WHERE rm.id = %s AND h.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p 
                WHERE p.id_usuario = %s
            )
        """, (registro_id, data['id_hilera'], usuario_id))
        
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"error": "No tienes acceso a este registro de mapeo o hilera"}), 403
        
        # Verificar si ya existe un estado para esta hilera en este registro
        cursor.execute("""
            SELECT id FROM mapeo_fact_estado_hilera 
            WHERE id_registro_mapeo = %s AND id_hilera = %s
        """, (registro_id, data['id_hilera']))
        
        estado_existente = cursor.fetchone()
        
        if estado_existente:
            # Actualizar estado existente
            cursor.execute("""
                UPDATE mapeo_fact_estado_hilera 
                SET estado = %s, fecha_actualizacion = NOW(), id_usuario = %s, observaciones = %s
                WHERE id = %s
            """, (
                data['estado'],
                usuario_id,
                data.get('observaciones'),
                estado_existente['id']
            ))
            estado_id = estado_existente['id']
        else:
            # Crear nuevo estado
            estado_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO mapeo_fact_estado_hilera (
                    id, id_registro_mapeo, id_hilera, estado, fecha_creacion, 
                    fecha_actualizacion, id_usuario, observaciones
                ) VALUES (%s, %s, %s, %s, NOW(), NOW(), %s, %s)
            """, (
                estado_id,
                registro_id,
                data['id_hilera'],
                data['estado'],
                usuario_id,
                data.get('observaciones')
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            "message": "Estado de hilera actualizado exitosamente",
            "id": estado_id
        }), 200
        
    except Exception as e:
        logger.error(f"Error creando/actualizando estado de hilera: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# ============================================================================
# REGISTROS INDIVIDUALES
# ============================================================================

# Obtener registros individuales
@mapeo_bp.route('/registros', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_registros():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener parámetros de filtro
        registro_mapeo_id = request.args.get('registro_mapeo_id')
        planta_id = request.args.get('planta_id')
        evaluador_id = request.args.get('evaluador_id')
        
        # Construir query base
        query = """
            SELECT r.*, p.planta, h.hilera, c.nombre as cuartel_nombre, 
                   u.nombre as evaluador_nombre, tp.nombre as tipo_planta_nombre
            FROM mapeo_fact_registro r
            JOIN general_dim_planta p ON r.id_planta = p.id
            JOIN general_dim_hilera h ON p.id_hilera = h.id
            JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_usuario u ON r.id_evaluador = u.id
            LEFT JOIN mapeo_dim_tipoplanta tp ON r.id_tipoplanta = tp.id
            WHERE c.id_sucursal IN (
                SELECT DISTINCT p2.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p2 
                WHERE p2.id_usuario = %s
            )
        """
        valores = [usuario_id]
        
        # Agregar filtros
        if registro_mapeo_id:
            query += " AND r.id_registro_mapeo = %s"
            valores.append(registro_mapeo_id)
        
        if planta_id:
            query += " AND r.id_planta = %s"
            valores.append(planta_id)
        
        if evaluador_id:
            query += " AND r.id_evaluador = %s"
            valores.append(evaluador_id)
        
        query += " ORDER BY r.hora_registro DESC"
        
        cursor.execute(query, valores)
        registros = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(registros), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo registros: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Crear un nuevo registro individual
@mapeo_bp.route('/registros', methods=['POST', 'OPTIONS'])
@jwt_required()
def crear_registro():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        # Validar datos requeridos
        campos_requeridos = ['id_planta', 'id_tipoplanta']
        for campo in campos_requeridos:
            if campo not in data:
                return jsonify({"error": f"Campo requerido: {campo}"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso a la planta
        cursor.execute("""
            SELECT 1 FROM general_dim_planta p
            JOIN general_dim_hilera h ON p.id_hilera = h.id
            JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            WHERE p.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p2.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p2 
                WHERE p2.id_usuario = %s
            )
        """, (data['id_planta'], usuario_id))
        
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({"error": "No tienes acceso a esta planta"}), 403
        
        # Generar UUID para el registro
        registro_id = str(uuid.uuid4())
        
        # Insertar el registro
        cursor.execute("""
            INSERT INTO mapeo_fact_registro (
                id, id_evaluador, hora_registro, id_planta, id_tipoplanta, 
                imagen, fecha_creacion
            ) VALUES (%s, %s, NOW(), %s, %s, %s, NOW())
        """, (
            registro_id,
            usuario_id,
            data['id_planta'],
            data['id_tipoplanta'],
            data.get('imagen')
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            "message": "Registro creado exitosamente",
            "id": registro_id
        }), 201
        
    except Exception as e:
        logger.error(f"Error creando registro: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Obtener un registro específico
@mapeo_bp.route('/registros/<string:registro_id>', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_registro(registro_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al registro
        cursor.execute("""
            SELECT r.*, p.planta, h.hilera, c.nombre as cuartel_nombre, 
                   u.nombre as evaluador_nombre, tp.nombre as tipo_planta_nombre
            FROM mapeo_fact_registro r
            JOIN general_dim_planta p ON r.id_planta = p.id
            JOIN general_dim_hilera h ON p.id_hilera = h.id
            JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_usuario u ON r.id_evaluador = u.id
            LEFT JOIN mapeo_dim_tipoplanta tp ON r.id_tipoplanta = tp.id
            WHERE r.id = %s AND c.id_sucursal IN (
                SELECT DISTINCT p2.id_sucursal 
                FROM usuario_pivot_sucursal_usuario p2 
                WHERE p2.id_usuario = %s
            )
        """, (registro_id, usuario_id))
        
        registro = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not registro:
            return jsonify({"error": "Registro no encontrado"}), 404
            
        return jsonify(registro), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo registro {registro_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# ============================================================================
# TIPOS DE PLANTA
# ============================================================================

# Obtener tipos de planta
@mapeo_bp.route('/tipos-planta', methods=['GET', 'OPTIONS'])
@jwt_required()
def obtener_tipos_planta():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener tipos de planta de la empresa del usuario
        cursor.execute("""
            SELECT tp.*
            FROM mapeo_dim_tipoplanta tp
            JOIN general_dim_usuario u ON tp.id_empresa = (
                SELECT DISTINCT e.id 
                FROM general_dim_empresa e
                JOIN general_dim_sucursal s ON e.id = s.id_empresa
                JOIN usuario_pivot_sucursal_usuario p ON s.id = p.id_sucursal
                WHERE p.id_usuario = %s
                LIMIT 1
            )
            WHERE tp.id_estado = 1
            ORDER BY tp.nombre ASC
        """, (usuario_id,))
        
        tipos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify(tipos), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo tipos de planta: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# ============================================================================
# CARGA MASIVA
# ============================================================================

def validar_coordenadas_gps(ubicacion):
    """Validar formato de coordenadas GPS (lat, lng)"""
    if not ubicacion:
        return True  # Ubicación es opcional
    
    try:
        # Formato esperado: "-33.123, -70.456"
        coords = ubicacion.split(',')
        if len(coords) != 2:
            return False
        
        lat = float(coords[0].strip())
        lng = float(coords[1].strip())
        
        # Validar rangos de latitud y longitud
        if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
            return False
        
        return True
    except (ValueError, AttributeError):
        return False

def validar_sucursal_usuario(cursor, usuario_id, id_sucursal):
    """Validar que el usuario tenga acceso a la sucursal"""
    cursor.execute("""
        SELECT 1 FROM usuario_pivot_sucursal_usuario 
        WHERE id_usuario = %s AND id_sucursal = %s
    """, (usuario_id, id_sucursal))
    return cursor.fetchone() is not None

def validar_sucursal_existe(cursor, id_sucursal):
    """Validar que la sucursal existe"""
    cursor.execute("""
        SELECT 1 FROM general_dim_sucursal 
        WHERE id = %s AND id_estado = 1
    """, (id_sucursal,))
    return cursor.fetchone() is not None

def validar_hilera_existe(cursor, id_hilera, id_cuartel):
    """Validar que la hilera existe y pertenece al cuartel"""
    cursor.execute("""
        SELECT 1 FROM general_dim_hilera 
        WHERE id = %s AND id_cuartel = %s AND id_estado = 1
    """, (id_hilera, id_cuartel))
    return cursor.fetchone() is not None

def validar_planta_existe(cursor, id_planta, id_hilera):
    """Validar que la planta existe y pertenece a la hilera"""
    cursor.execute("""
        SELECT 1 FROM general_dim_planta 
        WHERE id = %s AND id_hilera = %s AND id_estado = 1
    """, (id_planta, id_hilera))
    return cursor.fetchone() is not None

def validar_tipo_planta_existe(cursor, id_tipoplanta, usuario_id):
    """Validar que el tipo de planta existe y pertenece a la empresa del usuario"""
    cursor.execute("""
        SELECT 1 FROM mapeo_dim_tipoplanta tp
        JOIN general_dim_usuario u ON tp.id_empresa = (
            SELECT DISTINCT e.id 
            FROM general_dim_empresa e
            JOIN general_dim_sucursal s ON e.id = s.id_empresa
            JOIN usuario_pivot_sucursal_usuario p ON s.id = p.id_sucursal
            WHERE p.id_usuario = %s
            LIMIT 1
        )
        WHERE tp.id = %s AND tp.id_estado = 1
    """, (usuario_id, id_tipoplanta))
    return cursor.fetchone() is not None

# Carga masiva de cuarteles con hileras y plantas
@mapeo_bp.route('/cuarteles/bulk', methods=['POST', 'OPTIONS'])
@jwt_required()
def carga_masiva_cuarteles():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        if not data or 'cuarteles' not in data:
            return jsonify({"error": "Se requiere el campo 'cuarteles' con un array de cuarteles"}), 400
        
        cuarteles_data = data['cuarteles']
        
        # Validar límite de procesamiento
        if len(cuarteles_data) > 1000:
            return jsonify({"error": "Máximo 1000 cuarteles por carga masiva"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Iniciar transacción
        conn.begin()
        
        estadisticas = {
            "cuarteles_creados": 0,
            "hileras_creadas": 0,
            "plantas_creadas": 0,
            "errores": [],
            "warnings": []
        }
        
        try:
            for i, cuartel_data in enumerate(cuarteles_data):
                # Validar datos del cuartel
                campos_requeridos = ['nombre', 'id_sucursal', 'superficie', 'n_hileras']
                for campo in campos_requeridos:
                    if campo not in cuartel_data:
                        estadisticas["errores"].append({
                            "fila": i + 1,
                            "campo": campo,
                            "error": f"Campo requerido: {campo}"
                        })
                        continue
                
                # Validar que la sucursal existe
                if not validar_sucursal_existe(cursor, cuartel_data['id_sucursal']):
                    estadisticas["errores"].append({
                        "fila": i + 1,
                        "campo": "id_sucursal",
                        "error": f"Sucursal con ID {cuartel_data['id_sucursal']} no existe"
                    })
                    continue
                
                # Validar que el usuario tiene acceso a la sucursal
                if not validar_sucursal_usuario(cursor, usuario_id, cuartel_data['id_sucursal']):
                    estadisticas["errores"].append({
                        "fila": i + 1,
                        "campo": "id_sucursal",
                        "error": f"No tienes acceso a la sucursal {cuartel_data['id_sucursal']}"
                    })
                    continue
                
                # Crear cuartel
                cursor.execute("""
                    INSERT INTO general_dim_cuartel (
                        nombre, id_sucursal, superficie, id_variedad, ano_plantacion,
                        dsh, deh, n_hileras, id_estado, fecha_creacion
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """, (
                    cuartel_data['nombre'],
                    cuartel_data['id_sucursal'],
                    cuartel_data['superficie'],
                    cuartel_data.get('id_variedad'),
                    cuartel_data.get('ano_plantacion'),
                    cuartel_data.get('dsh'),
                    cuartel_data.get('deh'),
                    cuartel_data['n_hileras'],
                    1  # id_estado activo
                ))
                
                cuartel_id = cursor.lastrowid
                estadisticas["cuarteles_creados"] += 1
                
                # Generar hileras automáticamente según n_hileras
                n_hileras = cuartel_data['n_hileras']
                for j in range(1, n_hileras + 1):
                    # Crear hilera con numeración automática
                    cursor.execute("""
                        INSERT INTO general_dim_hilera (
                            hilera, id_cuartel, id_estado, fecha_creacion
                        ) VALUES (%s, %s, %s, NOW())
                    """, (
                        f"Hilera {j}",  # Hilera 1, Hilera 2, Hilera 3, etc.
                        cuartel_id,
                        1  # id_estado activo
                    ))
                    
                    hilera_id = cursor.lastrowid
                    estadisticas["hileras_creadas"] += 1
                    
                    # Si se especifican plantas para esta hilera, crearlas
                    if 'hileras' in cuartel_data and cuartel_data['hileras']:
                        # Buscar la hilera correspondiente en los datos
                        hilera_config = next((h for h in cuartel_data['hileras'] if h.get('hilera') == f"Hilera {j}"), None)
                        
                        if hilera_config and 'plantas' in hilera_config:
                            plantas_creadas_en_hilera = set()
                            
                            for k, planta_data in enumerate(hilera_config['plantas']):
                                if 'planta' not in planta_data:
                                    estadisticas["errores"].append({
                                        "fila": f"{i + 1}.{j}.{k + 1}",
                                        "campo": "planta",
                                        "error": "Campo requerido: planta"
                                    })
                                    continue
                                
                                # Validar coordenadas GPS si se proporcionan
                                if 'ubicacion' in planta_data:
                                    if not validar_coordenadas_gps(planta_data['ubicacion']):
                                        estadisticas["warnings"].append({
                                            "fila": f"{i + 1}.{j}.{k + 1}",
                                            "campo": "ubicacion",
                                            "warning": f"Formato de coordenadas inválido: {planta_data['ubicacion']}"
                                        })
                                        planta_data['ubicacion'] = None
                                
                                # Verificar duplicados en la hilera
                                if planta_data['planta'] in plantas_creadas_en_hilera:
                                    estadisticas["warnings"].append({
                                        "fila": f"{i + 1}.{j}.{k + 1}",
                                        "campo": "planta",
                                        "warning": f"Planta duplicada en Hilera {j}: Planta {planta_data['planta']}"
                                    })
                                    continue
                                
                                # Crear planta
                                cursor.execute("""
                                    INSERT INTO general_dim_planta (
                                        planta, id_hilera, ubicacion, id_estado, fecha_creacion
                                    ) VALUES (%s, %s, %s, %s, NOW())
                                """, (
                                    planta_data['planta'],
                                    hilera_id,
                                    planta_data.get('ubicacion'),
                                    1  # id_estado activo
                                ))
                                
                                plantas_creadas_en_hilera.add(planta_data['planta'])
                                estadisticas["plantas_creadas"] += 1
            
            # Commit de la transacción
            conn.commit()
            
            return jsonify({
                "success": True,
                "message": "Carga masiva de cuarteles completada exitosamente",
                "data": estadisticas
            }), 201
            
        except Exception as e:
            # Rollback en caso de error
            conn.rollback()
            logger.error(f"Error en carga masiva de cuarteles: {str(e)}")
            return jsonify({
                "success": False,
                "message": "Error en la carga masiva",
                "error": str(e)
            }), 500
        
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Error en carga masiva de cuarteles: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Carga masiva de registros de mapeo
@mapeo_bp.route('/registros/bulk', methods=['POST', 'OPTIONS'])
@jwt_required()
def carga_masiva_registros():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        if not data or 'registros' not in data:
            return jsonify({"error": "Se requiere el campo 'registros' con un array de registros"}), 400
        
        registros_data = data['registros']
        
        # Validar límite de procesamiento
        if len(registros_data) > 1000:
            return jsonify({"error": "Máximo 1000 registros por carga masiva"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Iniciar transacción
        conn.begin()
        
        estadisticas = {
            "registros_creados": 0,
            "errores": [],
            "warnings": []
        }
        
        try:
            for i, registro_data in enumerate(registros_data):
                # Validar datos del registro
                campos_requeridos = ['id_planta', 'id_tipoplanta']
                for campo in campos_requeridos:
                    if campo not in registro_data:
                        estadisticas["errores"].append({
                            "fila": i + 1,
                            "campo": campo,
                            "error": f"Campo requerido: {campo}"
                        })
                        continue
                
                # Validar que la planta existe y el usuario tiene acceso
                cursor.execute("""
                    SELECT p.id, h.id as id_hilera FROM general_dim_planta p
                    JOIN general_dim_hilera h ON p.id_hilera = h.id
                    JOIN general_dim_cuartel c ON h.id_cuartel = c.id
                    WHERE p.id = %s AND c.id_sucursal IN (
                        SELECT DISTINCT p2.id_sucursal 
                        FROM usuario_pivot_sucursal_usuario p2 
                        WHERE p2.id_usuario = %s
                    ) AND p.id_estado = 1
                """, (registro_data['id_planta'], usuario_id))
                
                planta_info = cursor.fetchone()
                if not planta_info:
                    estadisticas["errores"].append({
                        "fila": i + 1,
                        "campo": "id_planta",
                        "error": f"Planta con ID {registro_data['id_planta']} no existe o no tienes acceso"
                    })
                    continue
                
                # Validar que el tipo de planta existe
                if not validar_tipo_planta_existe(cursor, registro_data['id_tipoplanta'], usuario_id):
                    estadisticas["errores"].append({
                        "fila": i + 1,
                        "campo": "id_tipoplanta",
                        "error": f"Tipo de planta con ID {registro_data['id_tipoplanta']} no existe"
                    })
                    continue
                
                # Validar hora_registro si se proporciona
                hora_registro = None
                if 'hora_registro' in registro_data:
                    try:
                        hora_registro = datetime.strptime(registro_data['hora_registro'], '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        estadisticas["warnings"].append({
                            "fila": i + 1,
                            "campo": "hora_registro",
                            "warning": f"Formato de hora inválido: {registro_data['hora_registro']}. Usando hora actual."
                        })
                        hora_registro = None
                
                # Generar UUID para el registro
                registro_id = str(uuid.uuid4())
                
                # Crear registro
                cursor.execute("""
                    INSERT INTO mapeo_fact_registro (
                        id, id_evaluador, hora_registro, id_planta, id_tipoplanta, 
                        imagen, fecha_creacion
                    ) VALUES (%s, %s, %s, %s, %s, %s, NOW())
                """, (
                    registro_id,
                    registro_data.get('id_evaluador', usuario_id),
                    hora_registro or datetime.now(),
                    registro_data['id_planta'],
                    registro_data['id_tipoplanta'],
                    registro_data.get('imagen')
                ))
                
                estadisticas["registros_creados"] += 1
            
            # Commit de la transacción
            conn.commit()
            
            return jsonify({
                "success": True,
                "message": "Carga masiva de registros completada exitosamente",
                "data": estadisticas
            }), 201
            
        except Exception as e:
            # Rollback en caso de error
            conn.rollback()
            logger.error(f"Error en carga masiva de registros: {str(e)}")
            return jsonify({
                "success": False,
                "message": "Error en la carga masiva",
                "error": str(e)
            }), 500
        
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Error en carga masiva de registros: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Importar desde Excel/CSV
@mapeo_bp.route('/import/excel', methods=['POST', 'OPTIONS'])
@jwt_required()
def importar_excel():
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        
        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({"error": "No se envió ningún archivo"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No se seleccionó ningún archivo"}), 400
        
        # Verificar tipo de archivo
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({"error": "Solo se permiten archivos Excel (.xlsx, .xls) o CSV"}), 400
        
        tipo_importacion = request.form.get('tipo_importacion', 'completo')
        if tipo_importacion not in ['plantas', 'registros', 'completo']:
            return jsonify({"error": "Tipo de importación inválido. Valores válidos: plantas, registros, completo"}), 400
        
        # Aquí implementarías la lógica para procesar el archivo Excel/CSV
        # Por ahora retornamos un mensaje de que la funcionalidad está en desarrollo
        
        return jsonify({
            "success": True,
            "message": "Importación desde Excel/CSV en desarrollo",
            "data": {
                "archivo": file.filename,
                "tipo_importacion": tipo_importacion,
                "tamaño": len(file.read()),
                "nota": "Esta funcionalidad requiere implementación adicional con pandas/openpyxl"
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error en importación Excel: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# ============================================================================
# GESTIÓN INDEPENDIENTE DE HILERAS
# ============================================================================

# Agregar hileras a un cuartel existente
@mapeo_bp.route('/cuarteles/<int:cuartel_id>/agregar-hileras', methods=['POST', 'OPTIONS'])
@jwt_required()
def agregar_hileras_cuartel(cuartel_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        if not data or 'cantidad' not in data:
            return jsonify({"error": "Se requiere el campo 'cantidad' con el número de hileras a agregar"}), 400
        
        cantidad = data['cantidad']
        if not isinstance(cantidad, int) or cantidad <= 0:
            return jsonify({"error": "La cantidad debe ser un número entero positivo"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Iniciar transacción
        conn.begin()
        
        try:
            # Verificar que el usuario tenga acceso al cuartel
            cursor.execute("""
                SELECT c.*, COUNT(h.id) as hileras_actuales
                FROM general_dim_cuartel c
                LEFT JOIN general_dim_hilera h ON c.id = h.id_cuartel AND h.id_estado = 1
                WHERE c.id = %s AND c.id_sucursal IN (
                    SELECT DISTINCT p.id_sucursal 
                    FROM usuario_pivot_sucursal_usuario p 
                    WHERE p.id_usuario = %s
                ) AND c.id_estado = 1
                GROUP BY c.id
            """, (cuartel_id, usuario_id))
            
            cuartel_info = cursor.fetchone()
            if not cuartel_info:
                cursor.close()
                conn.close()
                return jsonify({"error": "Cuartel no encontrado o sin acceso"}), 404
            
            # Obtener el siguiente número de hilera
            siguiente_numero = cuartel_info['hileras_actuales'] + 1
            
            hileras_creadas = []
            
            # Crear las nuevas hileras
            for i in range(cantidad):
                numero_hilera = siguiente_numero + i
                
                cursor.execute("""
                    INSERT INTO general_dim_hilera (
                        hilera, id_cuartel, id_estado, fecha_creacion
                    ) VALUES (%s, %s, %s, NOW())
                """, (
                    f"Hilera {numero_hilera}",
                    cuartel_id,
                    1  # id_estado activo
                ))
                
                hilera_id = cursor.lastrowid
                hileras_creadas.append({
                    "id": hilera_id,
                    "hilera": f"Hilera {numero_hilera}"
                })
            
            # Actualizar el número de hileras en el cuartel
            nuevo_total = cuartel_info['hileras_actuales'] + cantidad
            cursor.execute("""
                UPDATE general_dim_cuartel 
                SET n_hileras = %s 
                WHERE id = %s
            """, (nuevo_total, cuartel_id))
            
            # Commit de la transacción
            conn.commit()
            
            return jsonify({
                "success": True,
                "message": f"Se agregaron {cantidad} hileras al cuartel exitosamente",
                "data": {
                    "cuartel_id": cuartel_id,
                    "hileras_agregadas": cantidad,
                    "total_hileras": nuevo_total,
                    "hileras_creadas": hileras_creadas
                }
            }), 201
            
        except Exception as e:
            # Rollback en caso de error
            conn.rollback()
            logger.error(f"Error agregando hileras al cuartel {cuartel_id}: {str(e)}")
            return jsonify({
                "success": False,
                "message": "Error agregando hileras",
                "error": str(e)
            }), 500
        
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Error agregando hileras al cuartel {cuartel_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Eliminar hilera con plantas asociadas (eliminación en cascada)
@mapeo_bp.route('/hileras/<int:hilera_id>', methods=['DELETE', 'OPTIONS'])
@jwt_required()
def eliminar_hilera_cascada(hilera_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Iniciar transacción
        conn.begin()
        
        try:
            # Verificar que el usuario tenga acceso a la hilera
            cursor.execute("""
                SELECT h.*, c.nombre as cuartel_nombre, c.n_hileras
                FROM general_dim_hilera h
                JOIN general_dim_cuartel c ON h.id_cuartel = c.id
                WHERE h.id = %s AND c.id_sucursal IN (
                    SELECT DISTINCT p.id_sucursal 
                    FROM usuario_pivot_sucursal_usuario p 
                    WHERE p.id_usuario = %s
                ) AND h.id_estado = 1
            """, (hilera_id, usuario_id))
            
            hilera_info = cursor.fetchone()
            if not hilera_info:
                cursor.close()
                conn.close()
                return jsonify({"error": "Hilera no encontrada o sin acceso"}), 404
            
            # Contar plantas asociadas a la hilera
            cursor.execute("""
                SELECT COUNT(*) as total_plantas
                FROM general_dim_planta 
                WHERE id_hilera = %s AND id_estado = 1
            """, (hilera_id,))
            
            plantas_count = cursor.fetchone()['total_plantas']
            
            # Eliminar plantas asociadas (soft delete)
            cursor.execute("""
                UPDATE general_dim_planta 
                SET id_estado = 0, fecha_baja = NOW()
                WHERE id_hilera = %s AND id_estado = 1
            """, (hilera_id,))
            
            # Eliminar la hilera (soft delete)
            cursor.execute("""
                UPDATE general_dim_hilera 
                SET id_estado = 0, fecha_baja = NOW()
                WHERE id = %s
            """, (hilera_id,))
            
            # Actualizar el número de hileras en el cuartel
            nuevo_total = hilera_info['n_hileras'] - 1
            cursor.execute("""
                UPDATE general_dim_cuartel 
                SET n_hileras = %s 
                WHERE id = %s
            """, (nuevo_total, hilera_info['id_cuartel']))
            
            # Commit de la transacción
            conn.commit()
            
            return jsonify({
                "success": True,
                "message": "Hilera eliminada exitosamente con todas sus plantas",
                "data": {
                    "hilera_id": hilera_id,
                    "hilera_nombre": hilera_info['hilera'],
                    "cuartel_id": hilera_info['id_cuartel'],
                    "cuartel_nombre": hilera_info['cuartel_nombre'],
                    "plantas_eliminadas": plantas_count,
                    "nuevo_total_hileras": nuevo_total
                }
            }), 200
            
        except Exception as e:
            # Rollback en caso de error
            conn.rollback()
            logger.error(f"Error eliminando hilera {hilera_id}: {str(e)}")
            return jsonify({
                "success": False,
                "message": "Error eliminando hilera",
                "error": str(e)
            }), 500
        
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Error eliminando hilera {hilera_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

# Actualizar estado de catastro de un cuartel
@mapeo_bp.route('/cuarteles/<int:cuartel_id>/estado-catastro', methods=['PUT', 'OPTIONS'])
@jwt_required()
def actualizar_estado_catastro(cuartel_id):
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        usuario_id = get_jwt_identity()
        data = request.get_json()
        
        if not data or 'estado_catastro' not in data:
            return jsonify({"error": "Se requiere el campo 'estado_catastro'"}), 400
        
        estado_catastro = data['estado_catastro']
        if estado_catastro not in ['pendiente', 'en_progreso', 'completado', 'verificado']:
            return jsonify({"error": "Estado de catastro inválido. Valores válidos: pendiente, en_progreso, completado, verificado"}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        try:
            # Verificar que el usuario tenga acceso al cuartel
            cursor.execute("""
                SELECT 1 FROM general_dim_cuartel c
                WHERE c.id = %s AND c.id_sucursal IN (
                    SELECT DISTINCT p.id_sucursal 
                    FROM usuario_pivot_sucursal_usuario p 
                    WHERE p.id_usuario = %s
                ) AND c.id_estado = 1
            """, (cuartel_id, usuario_id))
            
            if not cursor.fetchone():
                cursor.close()
                conn.close()
                return jsonify({"error": "Cuartel no encontrado o sin acceso"}), 404
            
            # Actualizar estado de catastro
            cursor.execute("""
                UPDATE general_dim_cuartel 
                SET estado_catastro = %s, fecha_actualizacion = NOW()
                WHERE id = %s
            """, (estado_catastro, cuartel_id))
            
            conn.commit()
            
            return jsonify({
                "success": True,
                "message": "Estado de catastro actualizado exitosamente",
                "data": {
                    "cuartel_id": cuartel_id,
                    "estado_catastro": estado_catastro
                }
            }), 200
            
        except Exception as e:
            conn.rollback()
            logger.error(f"Error actualizando estado de catastro del cuartel {cuartel_id}: {str(e)}")
            return jsonify({
                "success": False,
                "message": "Error actualizando estado de catastro",
                "error": str(e)
            }), 500
        
        finally:
            cursor.close()
            conn.close()
        
    except Exception as e:
        logger.error(f"Error actualizando estado de catastro del cuartel {cuartel_id}: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500

@mapeo_bp.route('/plantillas/<tipo>', methods=['GET'])
@jwt_required()
def descargar_plantilla_excel(tipo):
    """
    Descargar plantilla de Excel para carga masiva
    """
    try:
        # Validar tipo de plantilla
        tipos_validos = ['cuarteles', 'registros', 'completo']
        if tipo not in tipos_validos:
            return jsonify({
                'success': False,
                'message': f'Tipo de plantilla inválido. Tipos válidos: {", ".join(tipos_validos)}'
            }), 400
        
        # Crear workbook
        wb = Workbook()
        
        if tipo == 'cuarteles':
            crear_plantilla_cuarteles(wb)
        elif tipo == 'registros':
            crear_plantilla_registros(wb)
        elif tipo == 'completo':
            crear_plantilla_completa(wb)
        
        # Guardar archivo temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Enviar archivo
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'plantilla_carga_masiva_{tipo}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logging.error(f"Error generando plantilla Excel: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Error generando plantilla Excel'
        }), 500

def crear_plantilla_cuarteles(wb):
    """Crear plantilla para carga masiva de cuarteles"""
    ws = wb.active
    ws.title = "Carga Masiva Cuarteles"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers principales
    headers_principales = [
        'A1:Cuartel', 'B1:Nombre', 'C1:ID Sucursal', 'D1:Superficie', 
        'E1:Número de Hileras', 'F1:ID Variedad', 'G1:Año Plantación', 
        'H1:DSH', 'I1:DEH'
    ]
    
    for header in headers_principales:
        cell = ws[header.split(':')[0]]
        cell.value = header.split(':')[1]
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers de hileras y plantas
    headers_hileras = [
        'J1:Hilera', 'K1:Planta', 'L1:Ubicación GPS'
    ]
    
    for header in headers_hileras:
        cell = ws[header.split(':')[0]]
        cell.value = header.split(':')[1]
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ejemplos de datos
    ejemplos = [
        ['Cuartel A', 'Cuartel Principal', 1, 15.5, 10, 1, 2020, 3.5, 1.2, 'Hilera 1', 'Planta 1', '-33.123, -70.456'],
        ['', '', '', '', '', '', '', '', '', 'Hilera 1', 'Planta 2', '-33.124, -70.457'],
        ['', '', '', '', '', '', '', '', '', 'Hilera 2', 'Planta 1', '-33.125, -70.458'],
        ['Cuartel B', 'Cuartel Secundario', 1, 12.0, 8, 2, 2019, 3.0, 1.0, 'Hilera 1', 'Planta 1', '-33.126, -70.459']
    ]
    
    for row_idx, ejemplo in enumerate(ejemplos, start=2):
        for col_idx, valor in enumerate(ejemplo, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [15, 20, 12, 12, 18, 12, 15, 8, 8, 12, 12, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Agregar hoja de instrucciones
    ws_instrucciones = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE CUARTELES"],
        [""],
        ["1. DATOS DEL CUARTEL:"],
        ["   - Nombre: Nombre del cuartel (requerido)"],
        ["   - ID Sucursal: ID de la sucursal (requerido)"],
        ["   - Superficie: Superficie en hectáreas (requerido)"],
        ["   - Número de Hileras: Cantidad de hileras a generar automáticamente (requerido)"],
        ["   - ID Variedad: ID de la variedad (opcional)"],
        ["   - Año Plantación: Año de plantación (opcional)"],
        ["   - DSH: Distancia entre hileras (opcional)"],
        ["   - DEH: Distancia entre plantas (opcional)"],
        [""],
        ["2. DATOS DE HILERAS Y PLANTAS (OPCIONAL):"],
        ["   - Hilera: Nombre de la hilera (ej: 'Hilera 1', 'Hilera 2')"],
        ["   - Planta: Nombre de la planta (ej: 'Planta 1', 'Planta 2')"],
        ["   - Ubicación GPS: Coordenadas en formato 'lat, lng' (opcional)"],
        [""],
        ["3. REGLAS:"],
        ["   - Las hileras se generan automáticamente según 'Número de Hileras'"],
        ["   - Si especificas plantas, deben corresponder a hileras existentes"],
        ["   - Las coordenadas GPS deben estar en formato 'lat, lng'"],
        ["   - No se permiten plantas duplicadas en la misma hilera"],
        [""],
        ["4. EJEMPLO:"],
        ["   - Cuartel con 10 hileras: Se generan automáticamente Hilera 1 a Hilera 10"],
        ["   - Si especificas plantas para 'Hilera 1', se crearán en esa hilera"],
        ["   - Si no especificas plantas, solo se crean las hileras vacías"]
    ]
    
    for row_idx, instruccion in enumerate(instrucciones, start=1):
        cell = ws_instrucciones.cell(row=row_idx, column=1, value=instruccion[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif ":" in instruccion[0]:
            cell.font = Font(bold=True)
    
    ws_instrucciones.column_dimensions['A'].width = 80

def crear_plantilla_registros(wb):
    """Crear plantilla para carga masiva de registros"""
    ws = wb.active
    ws.title = "Carga Masiva Registros"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'A1:ID Planta', 'B1:ID Tipo Planta', 'C1:ID Evaluador', 
        'D1:Hora Registro', 'E1:Imagen (Base64)'
    ]
    
    for header in headers:
        cell = ws[header.split(':')[0]]
        cell.value = header.split(':')[1]
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ejemplos
    ejemplos = [
        [123, 'uuid-tipo-planta-1', 'user123', '2024-01-15 10:30:00', ''],
        [124, 'uuid-tipo-planta-2', 'user456', '2024-01-15 10:35:00', ''],
        [125, 'uuid-tipo-planta-1', 'user123', '2024-01-15 10:40:00', '']
    ]
    
    for row_idx, ejemplo in enumerate(ejemplos, start=2):
        for col_idx, valor in enumerate(ejemplo, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [12, 25, 15, 20, 50]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Agregar hoja de instrucciones
    ws_instrucciones = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE REGISTROS"],
        [""],
        ["1. DATOS REQUERIDOS:"],
        ["   - ID Planta: ID numérico de la planta (requerido)"],
        ["   - ID Tipo Planta: UUID del tipo de planta (requerido)"],
        [""],
        ["2. DATOS OPCIONALES:"],
        ["   - ID Evaluador: ID del evaluador"],
        ["   - Hora Registro: Fecha y hora en formato 'YYYY-MM-DD HH:MM:SS'"],
        ["   - Imagen: Datos de imagen en formato Base64"],
        [""],
        ["3. REGLAS:"],
        ["   - La planta debe existir en el sistema"],
        ["   - El tipo de planta debe ser válido"],
        ["   - La fecha debe estar en formato correcto"],
        ["   - La imagen es opcional"]
    ]
    
    for row_idx, instruccion in enumerate(instrucciones, start=1):
        cell = ws_instrucciones.cell(row=row_idx, column=1, value=instruccion[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif ":" in instruccion[0]:
            cell.font = Font(bold=True)
    
    ws_instrucciones.column_dimensions['A'].width = 80

def crear_plantilla_completa(wb):
    """Crear plantilla completa con todas las opciones"""
    # Crear hojas para cada tipo
    crear_plantilla_cuarteles(wb)
    crear_plantilla_registros(wb)
    
    # Renombrar hojas
    wb['Carga Masiva Cuarteles'].title = "Cuarteles"
    wb['Carga Masiva Registros'].title = "Registros"
    
    # Agregar hoja de índice
    ws_indice = wb.create_sheet("Índice", 0)
    ws_indice.title = "Índice"
    
    # Contenido del índice
    contenido_indice = [
        ["PLANTILLA COMPLETA - CARGA MASIVA"],
        [""],
        ["Esta plantilla contiene:"],
        ["1. Hoja 'Cuarteles': Para carga masiva de cuarteles con hileras y plantas"],
        ["2. Hoja 'Registros': Para carga masiva de registros de mapeo"],
        ["3. Hoja 'Instrucciones': Instrucciones detalladas para cada tipo"],
        [""],
        ["INSTRUCCIONES GENERALES:"],
        ["- Cada hoja tiene su propia estructura y validaciones"],
        ["- Sigue las instrucciones específicas de cada hoja"],
        ["- Los campos requeridos están marcados en las instrucciones"],
        ["- Ejemplos de datos están incluidos en cada hoja"],
        ["- No modifiques la estructura de las columnas"],
        [""],
        ["FORMATOS ACEPTADOS:"],
        ["- Fechas: YYYY-MM-DD HH:MM:SS"],
        ["- Coordenadas GPS: 'lat, lng' (ej: '-33.123, -70.456')"],
        ["- Imágenes: Base64 (opcional)"],
        ["- UUIDs: Formato estándar UUID"]
    ]
    
    for row_idx, contenido in enumerate(contenido_indice, start=1):
        cell = ws_indice.cell(row=row_idx, column=1, value=contenido[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_idx in [4, 5, 8, 13]:
            cell.font = Font(bold=True)
    
    ws_indice.column_dimensions['A'].width = 80
