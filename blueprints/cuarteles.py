from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
import logging
from utils.db import get_db_connection
from datetime import datetime

cuarteles_bp = Blueprint('cuarteles_bp', __name__)

# Configurar logging
logger = logging.getLogger(__name__)

@cuarteles_bp.route('/cuarteles', methods=['GET'])
@jwt_required()
def listar_cuarteles():
    """
    Listar todos los cuarteles del usuario logueado
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Obtener cuarteles del usuario según sus sucursales asignadas
        query = """
            SELECT 
                c.id,
                c.id_ceco,
                c.nombre,
                c.id_variedad,
                c.superficie,
                c.ano_plantacion,
                c.dsh,
                c.deh,
                c.id_propiedad,
                c.id_portainjerto,
                c.brazos_ejes,
                c.id_estado,
                c.fecha_baja,
                c.id_estadoproductivo,
                c.n_hileras,
                c.id_estadocatastro,
                s.nombre as nombre_sucursal,
                v.nombre as nombre_variedad
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            LEFT JOIN general_dim_variedad v ON c.id_variedad = v.id
            WHERE ce.id_sucursal IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            AND c.id_estado = 1
            ORDER BY c.nombre
        """
        
        cursor.execute(query, (user_id,))
        cuarteles = cursor.fetchall()
        
        # Log para debug
        logger.info(f"Usuario ID: {user_id}")
        logger.info(f"Cuarteles encontrados: {len(cuarteles)}")
        
        # Consulta adicional para debug - verificar cuarteles por sucursal
        debug_query = """
            SELECT 
                c.id,
                c.nombre,
                c.id_ceco,
                c.id_estado,
                s.nombre as sucursal_nombre
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            WHERE ce.id_sucursal IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            ORDER BY ce.id_sucursal, c.nombre
        """
        cursor.execute(debug_query, (user_id,))
        debug_info = cursor.fetchall()
        logger.info(f"Debug - Cuarteles por sucursal: {debug_info}")
        
        # Consulta para verificar todos los cuarteles
        all_cuarteles_query = """
            SELECT 
                c.id,
                c.nombre,
                c.id_ceco,
                c.id_estado,
                s.nombre as sucursal_nombre
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_sucursal s ON c.id_ceco = s.id
            ORDER BY c.id_ceco, c.nombre
            LIMIT 10
        """
        cursor.execute(all_cuarteles_query)
        all_cuarteles = cursor.fetchall()
        logger.info(f"Debug - Primeros 10 cuarteles en BD: {all_cuarteles}")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Cuarteles obtenidos exitosamente",
            "data": {
                "cuarteles": cuarteles,
                "total": len(cuarteles)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo cuarteles: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>', methods=['GET'])
@jwt_required()
def obtener_cuartel(cuartel_id):
    """
    Obtener un cuartel específico
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al cuartel
        query = """
            SELECT 
                c.id,
                c.id_ceco,
                c.nombre,
                c.id_variedad,
                c.superficie,
                c.ano_plantacion,
                c.dsh,
                c.deh,
                c.id_propiedad,
                c.id_portainjerto,
                c.brazos_ejes,
                c.id_estado,
                c.fecha_baja,
                c.id_estadoproductivo,
                c.n_hileras,
                c.id_estadocatastro,
                s.nombre as nombre_sucursal,
                v.nombre as nombre_variedad
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_sucursal s ON c.id_ceco = s.id
            LEFT JOIN general_dim_variedad v ON c.id_variedad = v.id
            INNER JOIN usuario_pivot_sucursal_usuario upsu ON s.id = upsu.id_sucursal
            WHERE c.id = %s 
            AND upsu.id_usuario = %s 
            AND c.id_estado = 1
        """
        
        cursor.execute(query, (cuartel_id, user_id))
        cuartel = cursor.fetchone()
        
        if not cuartel:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Cuartel no encontrado o sin permisos de acceso"
            }), 404
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Cuartel obtenido exitosamente",
            "data": {
                "cuartel": cuartel
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>', methods=['PUT'])
@jwt_required()
def editar_cuartel(cuartel_id):
    """
    Editar un cuartel existente
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener datos del request
        data = request.get_json()
        
        # Validar datos requeridos
        if not data or 'nombre' not in data:
            return jsonify({
                "success": False,
                "message": "El nombre del cuartel es requerido"
            }), 400
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al cuartel
        query_verificar = """
            SELECT c.id 
            FROM general_dim_cuartel c
            INNER JOIN usuario_pivot_sucursal_usuario upsu ON c.id_ceco = upsu.id_sucursal
            WHERE c.id = %s 
            AND upsu.id_usuario = %s 
            AND c.id_estado = 1
        """
        
        cursor.execute(query_verificar, (cuartel_id, user_id))
        cuartel_existe = cursor.fetchone()
        
        if not cuartel_existe:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Cuartel no encontrado o sin permisos de acceso"
            }), 404
        
        # Preparar datos para actualización
        campos_actualizables = [
            'nombre', 'id_variedad', 'superficie', 'ano_plantacion',
            'dsh', 'deh', 'id_propiedad', 'id_portainjerto', 'brazos_ejes',
            'id_estadoproductivo', 'n_hileras', 'id_estadocatastro'
        ]
        
        campos_a_actualizar = []
        valores = []
        
        for campo in campos_actualizables:
            if campo in data:
                campos_a_actualizar.append(f"{campo} = %s")
                valores.append(data[campo])
        
        if not campos_a_actualizar:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "No se proporcionaron campos para actualizar"
            }), 400
        
        # Construir query de actualización
        query_actualizar = f"""
            UPDATE general_dim_cuartel 
            SET {', '.join(campos_a_actualizar)}
            WHERE id = %s
        """
        valores.append(cuartel_id)
        
        cursor.execute(query_actualizar, valores)
        conn.commit()
        
        # Obtener cuartel actualizado
        query_obtener = """
            SELECT 
                c.id,
                c.id_ceco,
                c.nombre,
                c.id_variedad,
                c.superficie,
                c.ano_plantacion,
                c.dsh,
                c.deh,
                c.id_propiedad,
                c.id_portainjerto,
                c.brazos_ejes,
                c.id_estado,
                c.fecha_baja,
                c.id_estadoproductivo,
                c.n_hileras,
                c.id_estadocatastro,
                s.nombre as nombre_sucursal,
                v.nombre as nombre_variedad
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_sucursal s ON c.id_ceco = s.id
            LEFT JOIN general_dim_variedad v ON c.id_variedad = v.id
            WHERE c.id = %s
        """
        
        cursor.execute(query_obtener, (cuartel_id,))
        cuartel_actualizado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Cuartel actualizado exitosamente",
            "data": {
                "cuartel": cuartel_actualizado
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error actualizando cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>', methods=['DELETE'])
@jwt_required()
def eliminar_cuartel(cuartel_id):
    """
    Eliminar (desactivar) un cuartel
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario tenga acceso al cuartel
        query_verificar = """
            SELECT c.id, c.nombre
            FROM general_dim_cuartel c
            INNER JOIN usuario_pivot_sucursal_usuario upsu ON c.id_ceco = upsu.id_sucursal
            WHERE c.id = %s 
            AND upsu.id_usuario = %s 
            AND c.id_estado = 1
        """
        
        cursor.execute(query_verificar, (cuartel_id, user_id))
        cuartel = cursor.fetchone()
        
        if not cuartel:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Cuartel no encontrado o sin permisos de acceso"
            }), 404
        
        # Realizar soft delete (cambiar estado a 0)
        query_eliminar = """
            UPDATE general_dim_cuartel 
            SET id_estado = 0, fecha_baja = NOW()
            WHERE id = %s
        """
        
        cursor.execute(query_eliminar, (cuartel_id,))
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Cuartel eliminado exitosamente",
            "data": {
                "cuartel_id": cuartel_id,
                "nombre": cuartel['nombre'],
                "eliminado_en": datetime.now().isoformat()
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error eliminando cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>/hileras', methods=['GET'])
@jwt_required()
def obtener_hileras_cuartel(cuartel_id):
    """
    Obtener hileras de un cuartel específico
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar acceso al cuartel y obtener hileras
        query = """
            SELECT 
                h.id,
                h.hilera,
                h.id_cuartel
            FROM general_dim_hilera h
            INNER JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            WHERE c.id = %s 
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            ORDER BY h.hilera
        """
        
        cursor.execute(query, (cuartel_id, user_id))
        hileras = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Hileras obtenidas exitosamente",
            "data": {
                "cuartel_id": cuartel_id,
                "hileras": hileras,
                "total": len(hileras)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo hileras del cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>/plantas', methods=['GET'])
@jwt_required()
def obtener_plantas_cuartel(cuartel_id):
    """
    Obtener plantas de un cuartel específico
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar acceso al cuartel y obtener plantas
        query = """
            SELECT 
                p.id,
                p.planta,
                p.id_hilera,
                p.ubicacion,
                h.hilera as nombre_hilera
            FROM general_dim_planta p
            INNER JOIN general_dim_hilera h ON p.id_hilera = h.id
            INNER JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            WHERE c.id = %s 
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            ORDER BY h.hilera, p.planta
        """
        
        cursor.execute(query, (cuartel_id, user_id))
        plantas = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Plantas obtenidas exitosamente",
            "data": {
                "cuartel_id": cuartel_id,
                "plantas": plantas,
                "total": len(plantas)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo plantas del cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>/hileras/<int:hilera_id>/plantas', methods=['GET'])
@jwt_required()
def obtener_plantas_hilera(cuartel_id, hilera_id):
    """
    Obtener plantas de una hilera específica
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar acceso a la hilera y obtener plantas
        query = """
            SELECT 
                p.id,
                p.planta,
                p.id_hilera,
                p.ubicacion,
                h.hilera as nombre_hilera
            FROM general_dim_planta p
            INNER JOIN general_dim_hilera h ON p.id_hilera = h.id
            INNER JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            WHERE c.id = %s 
            AND h.id = %s
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            ORDER BY p.planta
        """
        
        cursor.execute(query, (cuartel_id, hilera_id, user_id))
        plantas = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Plantas obtenidas exitosamente",
            "data": {
                "cuartel_id": cuartel_id,
                "hilera_id": hilera_id,
                "plantas": plantas,
                "total": len(plantas)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo plantas de la hilera {hilera_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/catastro-masivo', methods=['POST'])
@jwt_required()
def catastro_masivo():
    """
    Crear hileras masivamente para múltiples cuarteles
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener datos del request
        data = request.get_json()
        
        if not data or 'cuarteles' not in data:
            return jsonify({
                "success": False,
                "message": "Se requiere el campo 'cuarteles' en el body"
            }), 400
        
        cuarteles_data = data['cuarteles']
        
        if not isinstance(cuarteles_data, list) or len(cuarteles_data) == 0:
            return jsonify({
                "success": False,
                "message": "El campo 'cuarteles' debe ser una lista no vacía"
            }), 400
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cuarteles_procesados = 0
        hileras_creadas = 0
        errores = []
        
        for cuartel_data in cuarteles_data:
            try:
                cuartel_id = cuartel_data.get('id')
                n_hileras = cuartel_data.get('n_hileras')
                
                if not cuartel_id or not n_hileras:
                    errores.append(f"Cuartel {cuartel_id}: Faltan campos requeridos (id, n_hileras)")
                    continue
                
                if n_hileras <= 0:
                    errores.append(f"Cuartel {cuartel_id}: n_hileras debe ser mayor a 0")
                    continue
                
                # Verificar que el cuartel existe y pertenece al usuario
                query_verificar = """
                    SELECT c.id, c.nombre, c.n_hileras
                    FROM general_dim_cuartel c
                    LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
                    LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
                    WHERE c.id = %s 
                    AND ce.id_sucursal IN (
                        SELECT id_sucursal 
                        FROM usuario_pivot_sucursal_usuario 
                        WHERE id_usuario = %s
                    )
                    AND c.id_estado = 1
                """
                
                cursor.execute(query_verificar, (cuartel_id, user_id))
                cuartel = cursor.fetchone()
                
                if not cuartel:
                    errores.append(f"Cuartel {cuartel_id}: No encontrado o sin permisos de acceso")
                    continue
                
                # Verificar si ya existen hileras para este cuartel
                query_hileras_existentes = """
                    SELECT COUNT(*) as total
                    FROM general_dim_hilera
                    WHERE id_cuartel = %s
                """
                cursor.execute(query_hileras_existentes, (cuartel_id,))
                hileras_existentes = cursor.fetchone()['total']
                
                if hileras_existentes > 0:
                    errores.append(f"Cuartel {cuartel_id}: Ya tiene {hileras_existentes} hileras existentes")
                    continue
                
                # Crear las hileras
                for i in range(1, n_hileras + 1):
                    query_crear_hilera = """
                        INSERT INTO general_dim_hilera (id_cuartel, hilera)
                        VALUES (%s, %s)
                    """
                    cursor.execute(query_crear_hilera, (cuartel_id, i))
                
                # Actualizar el campo n_hileras del cuartel
                query_actualizar_cuartel = """
                    UPDATE general_dim_cuartel 
                    SET n_hileras = %s
                    WHERE id = %s
                """
                cursor.execute(query_actualizar_cuartel, (n_hileras, cuartel_id))
                
                cuarteles_procesados += 1
                hileras_creadas += n_hileras
                
                logger.info(f"Hileras creadas para cuartel {cuartel_id}: {n_hileras} hileras")
                
            except Exception as e:
                error_msg = f"Error procesando cuartel {cuartel_data.get('id', 'N/A')}: {str(e)}"
                errores.append(error_msg)
                logger.error(error_msg)
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if cuarteles_procesados == 0:
            return jsonify({
                "success": False,
                "message": "No se pudo procesar ningún cuartel",
                "data": {
                    "errores": errores
                }
            }), 400
        
        return jsonify({
            "success": True,
            "message": f"Se crearon hileras para {cuarteles_procesados} cuarteles",
            "data": {
                "cuarteles_procesados": cuarteles_procesados,
                "hileras_creadas": hileras_creadas,
                "errores": errores if errores else []
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error en catastro masivo: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/plantas-masivo', methods=['POST'])
@jwt_required()
def plantas_masivo():
    """
    Crear plantas masivamente para múltiples hileras
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener datos del request
        data = request.get_json()
        
        # Log para debug
        logger.info(f"Plantas masivo - Datos recibidos: {data}")
        
        if not data or 'plantas' not in data:
            logger.error(f"Plantas masivo - Error: No se encontró el campo 'plantas' en los datos")
            return jsonify({
                "success": False,
                "message": "Se requiere el campo 'plantas' en el body"
            }), 400
        
        plantas_data = data['plantas']
        
        logger.info(f"Plantas masivo - Plantas data: {plantas_data}")
        logger.info(f"Plantas masivo - Tipo de plantas_data: {type(plantas_data)}")
        logger.info(f"Plantas masivo - Longitud: {len(plantas_data) if isinstance(plantas_data, list) else 'No es lista'}")
        
        if not isinstance(plantas_data, list) or len(plantas_data) == 0:
            logger.error(f"Plantas masivo - Error: plantas_data no es una lista válida")
            return jsonify({
                "success": False,
                "message": "El campo 'plantas' debe ser una lista no vacía"
            }), 400
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        hileras_procesadas = 0
        plantas_creadas = 0
        errores = []
        
        for planta_data in plantas_data:
            try:
                id_cuartel = planta_data.get('id_cuartel')
                id_hilera = planta_data.get('id_hilera')
                n_plantas = planta_data.get('n_plantas')
                
                if not id_cuartel or not id_hilera or not n_plantas:
                    errores.append(f"Hilera {id_hilera}: Faltan campos requeridos (id_cuartel, id_hilera, n_plantas)")
                    continue
                
                if n_plantas <= 0:
                    errores.append(f"Hilera {id_hilera}: n_plantas debe ser mayor a 0")
                    continue
                
                                                 # Verificar que la hilera existe y pertenece al usuario
                query_verificar = """
                    SELECT h.id, h.hilera, h.id_cuartel, c.nombre as nombre_cuartel
                    FROM general_dim_hilera h
                    INNER JOIN general_dim_cuartel c ON h.id_cuartel = c.id
                    LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
                    LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
                    WHERE h.id = %s 
                    AND h.id_cuartel = %s
                    AND ce.id_sucursal IN (
                        SELECT id_sucursal 
                        FROM usuario_pivot_sucursal_usuario 
                        WHERE id_usuario = %s
                    )
                """
                
                cursor.execute(query_verificar, (id_hilera, id_cuartel, user_id))
                hilera = cursor.fetchone()
                
                if not hilera:
                    errores.append(f"Hilera {id_hilera}: No encontrada o sin permisos de acceso")
                    continue
                
                # Verificar si ya existen plantas para esta hilera
                query_plantas_existentes = """
                    SELECT COUNT(*) as total
                    FROM general_dim_planta
                    WHERE id_hilera = %s
                """
                cursor.execute(query_plantas_existentes, (id_hilera,))
                plantas_existentes = cursor.fetchone()['total']
                
                if plantas_existentes > 0:
                    errores.append(f"Hilera {id_hilera}: Ya tiene {plantas_existentes} plantas existentes")
                    continue
                
                # Crear las plantas
                for i in range(1, n_plantas + 1):
                    query_crear_planta = """
                        INSERT INTO general_dim_planta (id_hilera, planta, fecha_creacion)
                        VALUES (%s, %s, NOW())
                    """
                    cursor.execute(query_crear_planta, (id_hilera, i))
                
                hileras_procesadas += 1
                plantas_creadas += n_plantas
                
                logger.info(f"Plantas creadas para hilera {id_hilera}: {n_plantas} plantas")
                
            except Exception as e:
                error_msg = f"Error procesando hilera {planta_data.get('id_hilera', 'N/A')}: {str(e)}"
                errores.append(error_msg)
                logger.error(error_msg)
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if hileras_procesadas == 0:
            return jsonify({
                "success": False,
                "message": "No se pudo procesar ninguna hilera",
                "data": {
                    "errores": errores
                }
            }), 400
        
        return jsonify({
            "success": True,
            "message": f"Se crearon plantas para {hileras_procesadas} hileras",
            "data": {
                "hileras_procesadas": hileras_procesadas,
                "plantas_creadas": plantas_creadas,
                "errores": errores if errores else []
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error en plantas masivo: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/plantas-masivo-info', methods=['POST'])
@jwt_required()
def obtener_info_plantas_masivo():
    """
    Obtener información de hileras y plantas para múltiples cuarteles en una sola petición
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener datos del request
        data = request.get_json()
        
        if not data or 'cuarteles' not in data:
            return jsonify({
                "success": False,
                "message": "Se requiere el campo 'cuarteles' en el body"
            }), 400
        
        cuarteles_ids = data['cuarteles']
        
        if not isinstance(cuarteles_ids, list) or len(cuarteles_ids) == 0:
            return jsonify({
                "success": False,
                "message": "El campo 'cuarteles' debe ser una lista no vacía"
            }), 400
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Consulta SQL optimizada: una sola consulta para todos los cuarteles
        query = """
            SELECT 
                c.id as cuartel_id,
                c.nombre as cuartel_nombre,
                h.id as hilera_id,
                h.hilera as numero_hilera,
                COALESCE(p.plantas_existentes, 0) as plantas_existentes
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_hilera h ON c.id = h.id_cuartel
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            LEFT JOIN (
                SELECT 
                    id_hilera,
                    COUNT(*) as plantas_existentes
                FROM general_dim_planta
                GROUP BY id_hilera
            ) p ON h.id = p.id_hilera
            WHERE c.id IN ({})
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            AND c.id_estado = 1
            ORDER BY c.id, h.hilera
        """.format(','.join(['%s'] * len(cuarteles_ids)))
        
        # Ejecutar consulta con todos los IDs de cuarteles
        params = cuarteles_ids + [user_id]
        cursor.execute(query, params)
        resultados = cursor.fetchall()
        
        # Organizar datos por cuartel
        cuarteles_info = {}
        for row in resultados:
            cuartel_id = row['cuartel_id']
            
            if cuartel_id not in cuarteles_info:
                cuarteles_info[cuartel_id] = {
                    "id": cuartel_id,
                    "nombre": row['cuartel_nombre'],
                    "hileras": []
                }
            
            if row['hilera_id']:  # Solo agregar si hay hilera
                cuarteles_info[cuartel_id]["hileras"].append({
                    "id": row['hilera_id'],
                    "hilera": row['numero_hilera'],
                    "plantas_existentes": row['plantas_existentes']
                })
        
        # Convertir a lista
        cuarteles_list = list(cuarteles_info.values())
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Información obtenida para {len(cuarteles_list)} cuarteles",
            "data": {
                "cuarteles": cuarteles_list,
                "total_cuarteles": len(cuarteles_list),
                "total_hileras": sum(len(c['hileras']) for c in cuarteles_list)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo info plantas masivo: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/<int:cuartel_id>/plantilla-plantas', methods=['GET'])
@jwt_required()
def descargar_plantilla_plantas(cuartel_id):
    """
    Descargar plantilla Excel para plantas de un cuartel específico
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verificar acceso al cuartel y obtener hileras con conteo de plantas
        query = """
            SELECT 
                h.id,
                h.hilera,
                h.id_cuartel,
                c.nombre as nombre_cuartel,
                COALESCE(p.plantas_existentes, 0) as plantas_existentes
            FROM general_dim_hilera h
            INNER JOIN general_dim_cuartel c ON h.id_cuartel = c.id
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            LEFT JOIN (
                SELECT 
                    id_hilera,
                    COUNT(*) as plantas_existentes
                FROM general_dim_planta
                GROUP BY id_hilera
            ) p ON h.id = p.id_hilera
            WHERE c.id = %s 
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            ORDER BY h.hilera
        """
        
        cursor.execute(query, (cuartel_id, user_id))
        hileras = cursor.fetchall()
        
        if not hileras:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "No se encontraron hileras para este cuartel"
            }), 404
        
        # Crear Excel con openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Plantas"
        
        # Estilos para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = ["ID_Cuartel", "Nombre_Cuartel", "ID_Hilera", "Nombre_Hilera", "Plantas_Existentes", "N_Plantas_Nuevas"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos con plantas existentes
        for row, hilera in enumerate(hileras, 2):
            ws.cell(row=row, column=1, value=hilera['id_cuartel'])
            ws.cell(row=row, column=2, value=hilera['nombre_cuartel'])
            ws.cell(row=row, column=3, value=hilera['id'])
            ws.cell(row=row, column=4, value=f"Hilera {hilera['hilera']}")
            ws.cell(row=row, column=5, value=hilera['plantas_existentes'])  # Plantas existentes
            ws.cell(row=row, column=6, value=0)  # Plantas nuevas a agregar
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        cursor.close()
        conn.close()
        
        # Crear respuesta con archivo
        from flask import send_file
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'plantilla_plantas_cuartel_{cuartel_id}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error descargando plantilla de plantas para cuartel {cuartel_id}: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500

@cuarteles_bp.route('/cuarteles/plantilla-plantas-masiva', methods=['POST'])
@jwt_required()
def descargar_plantilla_plantas_masiva():
    """
    Descargar plantilla Excel para múltiples cuarteles en una sola petición
    """
    try:
        # Obtener usuario logueado
        user_id = get_jwt_identity()
        
        # Obtener datos del request
        data = request.get_json()
        
        if not data or 'cuarteles' not in data:
            return jsonify({
                "success": False,
                "message": "Se requiere el campo 'cuarteles' en el body"
            }), 400
        
        cuarteles_ids = data['cuarteles']
        
        if not isinstance(cuarteles_ids, list) or len(cuarteles_ids) == 0:
            return jsonify({
                "success": False,
                "message": "El campo 'cuarteles' debe ser una lista no vacía"
            }), 400
        
        # Obtener conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Consulta SQL optimizada para múltiples cuarteles
        query = """
            SELECT 
                c.id as cuartel_id,
                c.nombre as cuartel_nombre,
                h.id as hilera_id,
                h.hilera as numero_hilera,
                COALESCE(p.plantas_existentes, 0) as plantas_existentes
            FROM general_dim_cuartel c
            LEFT JOIN general_dim_hilera h ON c.id = h.id_cuartel
            LEFT JOIN general_dim_ceco ce ON c.id_ceco = ce.id
            LEFT JOIN general_dim_sucursal s ON ce.id_sucursal = s.id
            LEFT JOIN (
                SELECT 
                    id_hilera,
                    COUNT(*) as plantas_existentes
                FROM general_dim_planta
                GROUP BY id_hilera
            ) p ON h.id = p.id_hilera
            WHERE c.id IN ({})
            AND s.id IN (
                SELECT id_sucursal 
                FROM usuario_pivot_sucursal_usuario 
                WHERE id_usuario = %s
            )
            AND c.id_estado = 1
            ORDER BY c.id, h.hilera
        """.format(','.join(['%s'] * len(cuarteles_ids)))
        
        # Ejecutar consulta
        params = cuarteles_ids + [user_id]
        cursor.execute(query, params)
        resultados = cursor.fetchall()
        
        if not resultados:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "No se encontraron cuarteles válidos"
            }), 404
        
        # Crear Excel con openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Plantas Masiva"
        
        # Estilos para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = ["ID_Cuartel", "Nombre_Cuartel", "ID_Hilera", "Nombre_Hilera", "Plantas_Existentes", "N_Plantas_Nuevas"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos de todos los cuarteles
        row = 2
        for resultado in resultados:
            ws.cell(row=row, column=1, value=resultado['cuartel_id'])
            ws.cell(row=row, column=2, value=resultado['cuartel_nombre'])
            ws.cell(row=row, column=3, value=resultado['hilera_id'])
            ws.cell(row=row, column=4, value=f"Hilera {resultado['numero_hilera']}")
            ws.cell(row=row, column=5, value=resultado['plantas_existentes'])
            ws.cell(row=row, column=6, value=0)  # Plantas nuevas a agregar
            row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        cursor.close()
        conn.close()
        
        # Crear respuesta con archivo
        from flask import send_file
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'plantilla_plantas_masiva_{len(cuarteles_ids)}_cuarteles.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error descargando plantilla masiva: {str(e)}")
        return jsonify({
            "success": False,
            "message": "Error interno del servidor"
        }), 500
